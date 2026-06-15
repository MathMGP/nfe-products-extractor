"""Escreve o xlsx de saída: cortes (formato Output) + dados de embarque
(container, booking, caixas, peso bruto) + TOTAIS/Média + coluna Status."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .core import CUT_NAMES

# Contrato, NF, Container, Booking, <8 cortes>, Peso Líquido, Peso Bruto, Caixas, Status
HEADER = (["Contrato", "NF", "Container", "Booking"] + CUT_NAMES
          + ["Peso Líquido", "Peso Bruto", "Caixas", "Status"])
N_COLS = len(HEADER)
# índices (1-based) úteis
C_CUT0 = 5            # primeira coluna de corte
C_NET = C_CUT0 + 8    # 13
C_GROSS = C_NET + 1   # 14
C_BOXES = C_NET + 2   # 15
C_STATUS = N_COLS     # 16
_NUMFMT = "#,##0.000"
_INTFMT = "#,##0"


def write_xlsx(results, out_path):
    """Gera o arquivo. Retorna (n_ok, n_problemas)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output"
    ws.append(HEADER)

    numeric = []
    for r in results:
        if r["vals"] is None:  # pasta não encontrada / sem NF / erro
            ws.append([r["contrato"], r["nf"], r["container"], r["booking"]]
                      + [None] * 8 + [None, None, None, r["status"]])
        else:
            ws.append([r["contrato"], r["nf"], r["container"], r["booking"]]
                      + r["vals"] + [r["net"], r["gross"], r["boxes"], r["status"]])
            numeric.append(r)

    n = len(numeric)
    if n:
        tot = ["TOTAIS", "", "", ""] + [
            round(sum(r["vals"][i] for r in numeric), 3) for i in range(8)
        ] + [
            round(sum(r["net"] for r in numeric), 3),
            round(sum(r["gross"] or 0 for r in numeric), 3),
            sum(r["boxes"] or 0 for r in numeric),
            "",
        ]
        ws.append(tot)
        avg = ["Média", "", "", ""] + [
            round(sum(r["vals"][i] for r in numeric) / n, 3) for i in range(8)
        ] + ["", "", "", ""]
        ws.append(avg)

    _style(ws, len(results), n)
    wb.save(out_path)
    return n, len(results) - n


def _style(ws, n_rows, n_numeric):
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="305496")
    warn_fill = PatternFill("solid", fgColor="FFE699")
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    last_data = n_rows + 1  # última linha de contrato
    # marca linhas com problema (sem valores numéricos) em amarelo
    for ri in range(2, last_data + 1):
        if ws.cell(ri, C_CUT0).value is None:
            for ci in range(1, N_COLS + 1):
                ws.cell(ri, ci).fill = warn_fill

    if n_numeric:  # negrito em TOTAIS/Média
        for ri in (last_data + 1, last_data + 2):
            for c in ws[ri]:
                c.font = Font(bold=True)

    # formatos: cortes + líquido + bruto = 3 casas; caixas = inteiro
    for col in range(C_CUT0, C_GROSS + 1):
        for row in range(2, last_data + 3):
            ws.cell(row, col).number_format = _NUMFMT
    for row in range(2, last_data + 3):
        ws.cell(row, C_BOXES).number_format = _INTFMT

    widths = ([10, 12, 13, 14] + [12] * 8 + [13, 13, 9, 40])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
